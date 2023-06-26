import cv2
import pytesseract
import os
import openpyxl
import re

# Set the path to the Tesseract OCR executable
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Get the current directory (where the script is located)
script_directory = os.getcwd()

# Set the directory containing the image files
image_directory = os.path.join(script_directory, 'images')

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active

# Set the column names in the first row of the Excel sheet
column_names = ['Date', 'Time', 'Longitude', 'Latitude']
for j, name in enumerate(column_names, start=1):
    column = openpyxl.utils.get_column_letter(j)
    sheet[column + '1'] = name

# Get a list of all image files in the directory
image_files = os.listdir(image_directory)

# Loop over each image file
for i, image_file in enumerate(image_files, start=2):  # Start from row 2 for data
    # Construct the full path to the image file
    image_path = os.path.join(image_directory, image_file)

    # Load the image
    image = cv2.imread(image_path)

    # Get the dimensions of the image
    height, width, _ = image.shape

    # Define the region of interest (top right corner)
    roi_width = int(width * 0.55)  # Adjust the width percentage as needed
    roi_height = int(height * 0.15)  # Adjust the height percentage as needed
    roi = image[0:roi_height, width - roi_width:width]

    # Apply any preprocessing steps here if needed

    # Perform OCR on the region of interest
    text = pytesseract.image_to_string(roi)

    # Define the regex patterns for extracting numbers ending with 'N' or 'E', time patterns in 'hh:mm' format, and date patterns in 'MM/dd/yy' format
    regex_patterns = [
        r'\d+/\d+/\d+',  # Date patterns in 'MM/dd/yy' format
        r'\d{2}:\d{2}',  # Time patterns in 'hh:mm' format
        r'\d+\.\d+N',  # Numbers ending with 'N'
        r'\d+\.\d+E'  # Numbers ending with 'E'
    ]

    # Extract numbers from each pattern and map them to respective columns
    data_row = []
    for pattern in regex_patterns:
        matches = re.findall(pattern, text)
        if matches:
            data_row.append(matches[0])
        else:
            data_row.append('')

    # Write the extracted data to the Excel sheet
    for j, data in enumerate(data_row, start=1):
        column = openpyxl.utils.get_column_letter(j)
        sheet[column + str(i)] = data

# Save the Excel workbook
workbook.save(os.path.join(script_directory, 'output.xlsx'))
